from openpyxl import Workbook
import os
from pymysql import* #pip install pymysql
import pandas.io.sql as sql #pip install pandas
import pandas as pd
from sqlalchemy import *
import sqlalchemy




def main(dbcon):
    a = 1
    print("-----------------------------------------------------------------")
    print("What would you like to do?\n"
          "0: Create project\n"
          "1: Update project\n"
          "2: Delete project\n"
          "3: Download project"
          "4: Update prices\n"
          "5: Quit"
          )
    awn = input(">>>")
    awn = awn.upper()
    while a == 1:
        if awn == '0':
            newproject(dbcon)
        elif awn == '1':
            get_table(dbcon,0)
        elif awn == '2':
            get_table(dbcon,1)
        elif awn == '3':
            get_table(dbcon,2)
        elif awn == '4':
            print("sorry this function is still WIP")
            main(dbcon)
        elif awn == '5':
            dbcon.close()
            input("Press any key to continue")
            quit()
        else:
            awn = input(">>>")

def newproject(dbcon):
    name = input("What is project name: ")
    name = name.lower()
    name = name.replace(" ","_")
    chars = set('0123456789$,"/?><:;!@#$%^&*()-=+[{]}|\\`~')
    if any((c in chars) for c in name):
        print('invalid name')
        print("")
        newproject(dbcon)
    else:
        locate_sheet(dbcon,name,0)

def create_table(dbcon,name,loc):
    cur = dbcon.cursor()
    sql ="""CREATE TABLE """ + name + """ (
   store_part_no CHAR(255) UNIQUE PRIMARY KEY NOT NULL,
   part_desc CHAR(255),
   quanity INT NOT NULL
    )"""
    cur.execute(sql)

    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    query = """INSERT INTO """ + name + """ (store_part_no,part_desc,quanity) VALUES (%s,%s,%s)"""
    for r in range(1,sheet.nrows):
        pfn = sheet.cell(r,0).value
        pd = sheet.cell(r,1).value
        num = sheet.cell(r,2).value

        values = (pfn,pd,num)
        valuesm = (pfn,pd)
        cur.execute(query, values)
        add_to_main(cur,valuesm)

    cur.close()
    db.commit()

    print("")
    print("All Done!")

    columns = str(sheet.ncols)
    rows = str(sheet.nrows)
    print("Commited "+columns+" columns and "+rows+" rows to database")

    #send back to main menu
    main(dbcon)

def locate_sheet(dbcon,table,flag):
    loc = os.getcwd()
    index = 0
    res = []

    print("Which sheet would you like to use? (select coresponding number): ")

    for file in os.listdir(loc):
    # check only text files
        if file.endswith('.xlsx'):
            print(str(index) + "> " + file)
            res.append(file)
            index = index + 1
    
    sheet = input(">>>")
    chars = set('$,"/?><:;!@#$%^&*()-=+[{]}|\\`~')
    
    while sheet.isalpha() or any((c in chars) for c in sheet):
        sheet = input(">>>")   

    while int(sheet) > index-1 or int(sheet) < 0:
        sheet = input(">>>")
    
    loc =  loc +"\\" + res[int(sheet)]

    if flag == 1:
        awn = input("This will overwrite all of main, do you wish to continue? Y/N: ")
        awn = awn.upper()
        if awn == 'Y':
           main_update(dbcon,table,loc)
        elif awn == 'N':
            get_table(dbcon)
    elif flag == 0:
        create_table(dbcon,table,loc)

    update(dbcon,table,loc)
    
def get_table(dbcon,flag):
    print("Which table would you like to change? ")
    
    insp = inspect(dbcon)
    tables = insp.get_table_names()

    print(tables)

    name = input(">>>")
    check = checkTableExists(dbcon,name)

    while check == False:
        name = input(">>>")
        check = checkTableExists(dbcon,name)

    if flag == 1:
        drop_table(dbcon,name)
    elif flag ==2:
        sql_to_excel(dbcon,name)
    if name == "main":
        locate_sheet(dbcon,name,1)

    print("")
    locate_sheet(dbcon,name,2)
        
def checkTableExists(dbcon,table):
   if sqlalchemy.inspect(dbcon).has_table(table):
       return True
   return False

def update(dbcon,table,loc):
    
    df = pd.read_excel(loc, engine = 'openpyxl')
    
    df.to_sql(table,dbcon,if_exists='replace')
    #df.to_sql('main',dbcon,if_exists='append')
    
    print("")
    print("All Done!")
    main(dbcon)

def main_update(dbcon,table,loc):
    cur = dbcon.cursor()

    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    
    cur.execute("TRUNCATE TABLE " + table)

    query = """INSERT INTO """ + table + """ (store_part_no,part_desc) VALUES (%s,%s)"""

    for r in range(1,sheet.nrows):
        pfn = sheet.cell(r,0).value
        pd = sheet.cell(r,1).value

        values = (pfn,pd)

        cur.execute(query, values)

    cur.close()
    dbcon.commit()

    print("")
    print("All Done!")

    columns = str(sheet.ncols)
    rows = str(sheet.nrows)
    print("Updated "+columns+" columns and "+rows+" rows to database")
    input("Press any key to continue")
    main(dbcon)

def drop_table(dbcon,name):
    cur = dbcon.cursor()
    if name == "main":
        print("Not authorized to delete main table")
        main(dbcon)
    
    awn = input("Are you sure you wish to delete the " + name + " project? Y/N: ")
    awn = awn.upper()

    while awn != 'Y' and awn != 'N':
        awn = input(">>>")
        awn = awn.upper()

    if awn == 'Y':
        awn = input("Are you 100 percent sure? Y/N: ")
        awn = awn.upper()

        while awn != 'Y' and awn != 'N':
            awn = input(">>>")
            awn = awn.upper()

        if awn == 'Y':
            cur.execute("DROP TABLE "+ name)
            print(name + "has been deleted")

    
    main(dbcon)

def sql_to_excel(dbcon,table):
    df=sql.read_sql('select * from '+table,dbcon)
    # export the data into the excel sheet
    df.to_excel(table+'.xlsx')
    main(dbcon)
def connect():
    engine = create_engine("mysql+pymysql://{0}:{1}@{2}/{3}".format(
            'root', 'SQLDatabasePassword228!','localhost','keyboardpricing'))
    return engine



db = connect()
main(db)
